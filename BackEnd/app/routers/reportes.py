"""Rutas para generar y descargar reportes Excel."""

import re
from datetime import date, datetime
from pathlib import Path
from uuid import uuid4

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    status,
)
from fastapi.responses import FileResponse
from openpyxl import Workbook
<<<<<<< HEAD
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
=======
>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import obtener_usuario_actual
from app.models.activo import Activo
from app.models.alerta import Alerta
from app.models.incidente import Incidente
from app.models.mantenimiento import Mantenimiento
from app.models.reporte import Reporte
from app.models.rol import Rol
from app.models.usuario import Usuario
from app.schemas.reporte import (
    ReporteCreate,
    ReporteResponse,
)


router = APIRouter(
    prefix="/reportes",
    tags=["Reportes"],
)

ROLES_REPORTES = {
    "ADMIN_EMPRESA",
    "AUDITOR",
}

DIRECTORIO_REPORTES = Path(
    "reportes_generados",
).resolve()

TIPO_EXCEL = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

CONFIGURACION_REPORTES = {
    "activos": (
        Activo,
        (
            "id",
            "sede_id",
            "tipo_activo_id",
            "nombre",
            "codigo_interno",
            "direccion_ip",
            "nombre_host",
            "sistema_operativo",
            "fabricante",
            "modelo",
            "numero_serie",
            "criticidad",
            "estado",
            "fecha_adquisicion",
            "descripcion",
            "creado_en",
        ),
    ),
    "alertas": (
        Alerta,
        (
            "id",
            "activo_id",
            "titulo",
            "descripcion",
            "severidad",
            "estado",
            "detectada_en",
            "resuelta_en",
        ),
    ),
    "incidentes": (
        Incidente,
        (
            "id",
            "activo_id",
            "asignado_a",
            "codigo",
            "titulo",
            "descripcion",
            "prioridad",
            "estado",
            "solucion",
            "abierto_en",
            "resuelto_en",
            "cerrado_en",
        ),
    ),
    "mantenimientos": (
        Mantenimiento,
        (
            "id",
            "activo_id",
            "responsable_id",
            "tipo",
            "estado",
            "descripcion",
            "resultado",
            "costo",
            "programado_para",
            "iniciado_en",
            "finalizado_en",
        ),
    ),
}


def obtener_usuario_reportes(
    usuario: Usuario = Depends(
        obtener_usuario_actual,
    ),
    db: Session = Depends(get_db),
) -> Usuario:
    rol = db.get(Rol, usuario.rol_id)

    if (
        rol is None
        or rol.nombre not in ROLES_REPORTES
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Se requiere el rol "
                "ADMIN_EMPRESA o AUDITOR."
            ),
        )

    return usuario


def obtener_reporte_de_la_organizacion(
    reporte_id: int,
    usuario: Usuario,
    db: Session,
) -> Reporte:
    reporte = db.scalar(
        select(Reporte).where(
            Reporte.id == reporte_id,
            Reporte.organizacion_id
            == usuario.organizacion_id,
        )
    )

    if reporte is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Reporte no encontrado.",
        )

    return reporte


def convertir_celda(valor: object) -> str:
    if valor is None:
        return ""

    if isinstance(valor, (date, datetime)):
        texto = valor.isoformat()
    else:
        texto = str(valor)

    texto = re.sub(
        r"[\x00-\x08\x0B-\x0C\x0E-\x1F]",
        "",
        texto,
    )

    if texto.startswith(
        ("=", "+", "-", "@", "\t", "\r")
    ):
        return f"'{texto}"

    return texto


def escribir_excel(
    tipo: str,
    organizacion_id: int,
    ruta_archivo: Path,
    db: Session,
) -> None:
    modelo, columnas = CONFIGURACION_REPORTES[tipo]

    consulta = (
        select(modelo)
        .where(
            modelo.organizacion_id
            == organizacion_id
        )
        .order_by(modelo.id)
    )

    registros = db.scalars(consulta).all()

    libro = Workbook()
    hoja = libro.active
    hoja.title = tipo.capitalize()
<<<<<<< HEAD
    hoja.append(list(columnas))

    for registro in registros:
        hoja.append(
            [
                convertir_celda(getattr(registro, campo))
=======

    hoja.append(list(columnas))

    for registro in registros:
        hoja.append(
            [
                convertir_celda(
                    getattr(registro, campo),
                )
>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)
                for campo in columnas
            ]
        )

<<<<<<< HEAD
    relleno_encabezado = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for celda in hoja[1]:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal="center")

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    for indice, columna in enumerate(
        hoja.iter_cols(),
        start=1,
    ):
        ancho = max(
            len(str(celda.value)) if celda.value is not None else 0
            for celda in columna
        )
        hoja.column_dimensions[
            get_column_letter(indice)
        ].width = min(ancho + 2, 45)
=======
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    for columna in hoja.columns:
        longitud = max(
            len(str(celda.value or ""))
            for celda in columna
        )

        hoja.column_dimensions[
            columna[0].column_letter
        ].width = min(longitud + 2, 50)
>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)

    libro.save(ruta_archivo)


@router.post(
    "",
    response_model=ReporteResponse,
    status_code=status.HTTP_201_CREATED,
)
def generar_reporte(
    datos: ReporteCreate,
    usuario: Usuario = Depends(
        obtener_usuario_reportes,
    ),
    db: Session = Depends(get_db),
) -> Reporte:
    directorio_organizacion = (
        DIRECTORIO_REPORTES
        / str(usuario.organizacion_id)
    )

    ruta_archivo = (
<<<<<<< HEAD
        directorio_organizacion / f"{uuid4().hex}.xlsx"
=======
        directorio_organizacion
        / f"{uuid4().hex}.xlsx"
>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)
    )

    try:
        directorio_organizacion.mkdir(
            parents=True,
            exist_ok=True,
        )
<<<<<<< HEAD
=======

>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)
        escribir_excel(
            datos.tipo,
            usuario.organizacion_id,
            ruta_archivo,
            db,
        )
    except OSError as error:
        raise HTTPException(
            status_code=(
                status.HTTP_500_INTERNAL_SERVER_ERROR
            ),
            detail=(
                "No fue posible generar "
                "el archivo Excel."
            ),
        ) from error

    reporte = Reporte(
        organizacion_id=usuario.organizacion_id,
        generado_por=usuario.id,
        nombre=datos.nombre,
        tipo=datos.tipo,
        formato="XLSX",
        ruta_archivo=str(ruta_archivo),
        parametros={},
    )

    try:
        db.add(reporte)
        db.commit()
        db.refresh(reporte)
    except SQLAlchemyError as error:
        db.rollback()
        ruta_archivo.unlink(missing_ok=True)

        raise HTTPException(
            status_code=(
                status.HTTP_500_INTERNAL_SERVER_ERROR
            ),
            detail=(
                "No fue posible guardar el reporte."
            ),
        ) from error

    return reporte


@router.get(
    "",
    response_model=list[ReporteResponse],
)
def listar_reportes(
    offset: int = Query(
        default=0,
        ge=0,
    ),
    limite: int = Query(
        default=50,
        ge=1,
        le=100,
    ),
    usuario: Usuario = Depends(
        obtener_usuario_reportes,
    ),
    db: Session = Depends(get_db),
) -> list[Reporte]:
    consulta = (
        select(Reporte)
        .where(
            Reporte.organizacion_id
            == usuario.organizacion_id
        )
        .order_by(
            Reporte.generado_en.desc(),
            Reporte.id.desc(),
        )
        .offset(offset)
        .limit(limite)
    )

    return list(
        db.scalars(consulta).all()
    )


@router.get("/{reporte_id}/descargar")
def descargar_reporte(
    reporte_id: int,
    usuario: Usuario = Depends(
        obtener_usuario_reportes,
    ),
    db: Session = Depends(get_db),
) -> FileResponse:
    reporte = obtener_reporte_de_la_organizacion(
        reporte_id,
        usuario,
        db,
    )

    ruta_archivo = Path(
        reporte.ruta_archivo,
    ).resolve()

    if (
        not ruta_archivo.is_relative_to(
            DIRECTORIO_REPORTES
        )
        or not ruta_archivo.is_file()
    ):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "Archivo de reporte no encontrado."
            ),
        )

    extension = ruta_archivo.suffix.lower()

    if extension == ".xlsx":
        media_type = TIPO_EXCEL
    elif extension == ".csv":
        media_type = "text/csv"
    else:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "Formato de reporte no permitido."
            ),
        )

    nombre_seguro = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        reporte.nombre,
    ).strip("_")

    nombre_descarga = (
<<<<<<< HEAD
        f"{nombre_seguro or 'reporte'}.xlsx"
=======
        f"{nombre_seguro or 'reporte'}"
        f"{extension}"
>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)
    )

    return FileResponse(
        path=ruta_archivo,
<<<<<<< HEAD
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
=======
        media_type=media_type,
>>>>>>> ac735ae (Actualizar reportes y configuracion del backend)
        filename=nombre_descarga,
    )
